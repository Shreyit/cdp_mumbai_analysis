"""
================================================================================
  Sustainability Report Text Extractor  —  Universal GRI Mapping Tool
================================================================================
  Loads expected GRI disclosure codes from gri_repo/standards/*.json corpus.
  Matches codes against company PDF text (code-only, no keywords).
  Extracts internal PDF hyperlinks to resolve actual reporting page numbers.
  Applies sector-specific disclosures only when company.sector_standard matches.

  Usage:
      python sustainability_extractor.py

  Output:
      output/<Company>_<Year>.json
      output/<Company>_<Year>.xlsx  (single GRI Coverage sheet)
================================================================================
"""

from __future__ import annotations

import os
import re
import json
import datetime
import pdfplumber
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 1 — CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

_ROOT     = str(Path(__file__).parent.parent)
_REPO_DIR = Path(_ROOT) / "gri_repo" / "standards"

CONFIG = {
    "company_name":    "Sony Group Corporation",
    "report_year":     "2025",
    "report_type":     "Sustainability Report",
    "sector_standard": None,   # "GRI 14" for mining, "GRI 11" for oil/gas, etc.
    "pdf_path": (
        _ROOT + "/pdfs/companies/"
        "https::www.sony.com:en:SonyInfo:csr:library:reports:SustainabilityReport2025_E.pdf"
    ),
    "output_dir": _ROOT + "/output",
}


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 2 — GRI TOPIC DEFINITIONS
#
#  standard_ids       → GRI topic standard IDs; disclosure codes are loaded
#                        at runtime from gri_repo/standards/*.json
#  sector_disclosures → sector-specific codes keyed by sector standard ID
#                        only applied when company.sector_standard matches
# ══════════════════════════════════════════════════════════════════════════════

GRI_TOPICS = {

    # ── Environment ───────────────────────────────────────────────────────────
    "Climate Change": {
        "standard_ids": ["GRI 302", "GRI 305"],
        "sector_disclosures": {},
    },
    "Air Emissions": {
        "standard_ids": ["GRI 305"],
        "sector_disclosures": {},
    },
    "Materials": {
        "standard_ids": ["GRI 301"],
        "sector_disclosures": {},
    },
    "Biodiversity": {
        "standard_ids": ["GRI 101", "GRI 304"],
        "sector_disclosures": {},
    },
    "Waste": {
        "standard_ids": ["GRI 306"],
        "sector_disclosures": {},
    },
    "Water and Effluents": {
        "standard_ids": ["GRI 303"],
        "sector_disclosures": {},
    },

    # ── Mining sector (GRI 14) ────────────────────────────────────────────────
    "Tailings": {
        "standard_ids": [],
        "sector_disclosures": {
            "GRI 14": {
                "14.6.2": "tailings disposal methods",
                "14.6.3": "tailings facilities",
            },
        },
    },
    "Closure and Rehabilitation": {
        "standard_ids": ["GRI 402", "GRI 404"],
        "sector_disclosures": {
            "GRI 14": {
                "14.8.4": "community transition plan",
                "14.8.5": "financial provisions for closure",
            },
        },
    },
    "Land and Resource Rights": {
        "standard_ids": [],
        "sector_disclosures": {
            "GRI 14": {
                "14.12.2": "involuntary resettlement",
            },
        },
    },
    "Artisanal and Small-Scale Mining": {
        "standard_ids": [],
        "sector_disclosures": {
            "GRI 14": {
                "14.13.2": "ASM engagement approaches",
                "14.13.3": "ASM support programs",
            },
        },
    },
    "Critical Incident Management": {
        "standard_ids": ["GRI 306"],
        "sector_disclosures": {
            "GRI 14": {
                "14.15.3": "critical incidents",
                "14.15.4": "emergency preparedness",
            },
        },
    },
    "Conflict-affected Areas": {
        "standard_ids": [],
        "sector_disclosures": {
            "GRI 14": {
                "14.25.2": "security and human rights in conflict areas",
            },
        },
    },

    # ── Oil & Gas (GRI 11) + Coal (GRI 12) ────────────────────────────────────
    "Decommissioning and Abandonment": {
        "standard_ids": [],
        "sector_disclosures": {
            "GRI 11": {
                "11.3.4": "decommissioning and abandonment liabilities",
            },
            "GRI 12": {
                "12.3.4": "decommissioning and abandonment liabilities",
            },
        },
    },
    "Asset Integrity and Process Safety": {
        "standard_ids": [],
        "sector_disclosures": {
            "GRI 11": {
                "11.2.4": "asset integrity and process safety",
            },
            "GRI 12": {
                "12.2.4": "asset integrity and process safety",
            },
        },
    },

    # ── Economic ──────────────────────────────────────────────────────────────
    "Economic Impacts": {
        "standard_ids": ["GRI 201", "GRI 203", "GRI 204"],
        "sector_disclosures": {},
    },
    "Payments to Governments": {
        "standard_ids": ["GRI 201", "GRI 207"],
        "sector_disclosures": {},
    },
    "Anti-competitive Behavior": {
        "standard_ids": ["GRI 206"],
        "sector_disclosures": {},
    },

    # ── Community & Human Rights ──────────────────────────────────────────────
    "Local Communities": {
        "standard_ids": ["GRI 413"],
        "sector_disclosures": {},
    },
    "Human Rights": {
        "standard_ids": ["GRI 411", "GRI 412"],
        "sector_disclosures": {},
    },
    "Security Practices": {
        "standard_ids": ["GRI 410"],
        "sector_disclosures": {},
    },

    # ── Labour ────────────────────────────────────────────────────────────────
    "Occupational Health & Safety": {
        "standard_ids": ["GRI 403"],
        "sector_disclosures": {},
    },
    "Employment Practices": {
        "standard_ids": ["GRI 202", "GRI 401", "GRI 402", "GRI 404", "GRI 414"],
        "sector_disclosures": {},
    },
    "Child Labour": {
        "standard_ids": ["GRI 408", "GRI 414"],
        "sector_disclosures": {},
    },
    "Forced Labour & Modern Slavery": {
        "standard_ids": ["GRI 409", "GRI 414"],
        "sector_disclosures": {},
    },
    "Freedom of Association": {
        "standard_ids": ["GRI 407"],
        "sector_disclosures": {},
    },
    "Non-discrimination & Equal Opportunity": {
        "standard_ids": ["GRI 202", "GRI 405", "GRI 406"],
        "sector_disclosures": {},
    },

    # ── Governance ────────────────────────────────────────────────────────────
    "Anti-corruption": {
        "standard_ids": ["GRI 205"],
        "sector_disclosures": {},
    },
    "Public Policy": {
        "standard_ids": ["GRI 415"],
        "sector_disclosures": {},
    },

    # ── Supply Chain & Product ────────────────────────────────────────────────
    "Responsible Supply Chain": {
        "standard_ids": ["GRI 308", "GRI 414"],
        "sector_disclosures": {},
    },
    "Product Safety & Quality": {
        "standard_ids": ["GRI 416"],
        "sector_disclosures": {},
    },
    "Marketing and Labeling": {
        "standard_ids": ["GRI 417"],
        "sector_disclosures": {},
    },
    "Data Privacy & Cybersecurity": {
        "standard_ids": ["GRI 418"],
        "sector_disclosures": {},
    },

    # ── Agriculture / Aquaculture / Fishing (GRI 13) ──────────────────────────
    "Food Security": {
        "standard_ids": [],
        "sector_disclosures": {
            "GRI 13": {
                "13.6.4": "impacts on food security",
                "13.6.5": "approaches to contributing to food security",
            },
        },
    },
    "Animal Welfare": {
        "standard_ids": [],
        "sector_disclosures": {
            "GRI 13": {
                "13.7.4": "approaches to animal welfare",
                "13.7.5": "animal welfare incidents",
            },
        },
    },
    "Pesticides and Agrochemicals": {
        "standard_ids": [],
        "sector_disclosures": {
            "GRI 13": {
                "13.8.4": "approaches to pesticide management",
                "13.8.5": "pesticides used",
            },
        },
    },
    "Fishing and Aquaculture Practices": {
        "standard_ids": [],
        "sector_disclosures": {
            "GRI 13": {
                "13.9.4": "fishing and aquaculture practices",
                "13.9.5": "impacts on marine ecosystems",
            },
        },
    },

    # ── Corporate Governance ──────────────────────────────────────────────────
    "Corporate Governance": {
        "standard_ids": ["GRI 2"],
        "sector_disclosures": {},
    },
}


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 3 — LOAD GRI STANDARDS FROM JSON CORPUS
# ══════════════════════════════════════════════════════════════════════════════

def load_gri_standards(repo_dir: Path) -> dict:
    """
    Load all GRI standard JSONs from gri_repo/standards/.
    Returns {standard_id: {..., disclosures: [{code, title, ...}]}}
    """
    standards: dict = {}
    if not repo_dir.exists():
        print(f"  [WARN] Standards repo not found: {repo_dir}")
        return standards
    for f in repo_dir.glob("*.json"):
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
            sid  = data.get("standard_id", "")
            if sid:
                standards[sid] = data
        except Exception:
            pass
    print(f"  [STANDARDS] Loaded {len(standards)} GRI standards from corpus")
    return standards


def get_topic_disclosures(topic_cfg: dict, standards: dict, sector_std: str | None) -> dict:
    """
    Returns {code: label} for all expected disclosures for this topic.
    - Topic standard codes come from the loaded JSON corpus (gri_repo/standards/)
    - Sector-specific codes are added only when sector_std matches
    """
    codes: dict[str, str] = {}

    for sid in topic_cfg.get("standard_ids", []):
        std = standards.get(sid)
        if std:
            for disc in std.get("disclosures", []):
                codes[disc["code"]] = disc.get("title", "")

    if sector_std:
        sector_discs = topic_cfg.get("sector_disclosures", {}).get(sector_std, {})
        codes.update(sector_discs)

    return codes


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 4 — PDF EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════

def extract_text_by_page(pdf_path: str) -> dict:
    """
    Extract text from every page.
    Returns {page_num: {"text": str, "combined": str}}.
    PyMuPDF primary (handles Asian-layout PDFs correctly), pdfplumber fallback.
    """
    pages: dict = {}

    try:
        import fitz
        doc   = fitz.open(pdf_path)
        total = len(doc)
        print(f"  [PDF] {total} pages (PyMuPDF)")
        for i in range(total):
            n    = i + 1
            text = doc[i].get_text().strip()
            pages[n] = {"text": text, "combined": text}
            if n % 50 == 0:
                print(f"  [PDF] {n}/{total} pages processed")
        doc.close()
        return pages
    except Exception as e:
        print(f"  [WARN] PyMuPDF error: {e} — trying pdfplumber")

    try:
        with pdfplumber.open(pdf_path) as pdf:
            total = len(pdf.pages)
            print(f"  [PDF] {total} pages (pdfplumber)")
            for i, page in enumerate(pdf.pages):
                n   = i + 1
                raw = page.extract_text() or ""
                tbl_lines = []
                for tbl in (page.extract_tables() or []):
                    for row in tbl:
                        if row:
                            tbl_lines.append(" | ".join(str(c) if c else "" for c in row))
                combined = f"{raw}\n{chr(10).join(tbl_lines)}".strip()
                pages[n] = {"text": raw.strip(), "combined": combined}
                if n % 50 == 0:
                    print(f"  [PDF] {n}/{total} pages processed")
        return pages
    except Exception as e2:
        print(f"  [ERROR] pdfplumber also failed: {e2}")
        return {}


def extract_hyperlinks_by_page(pdf_path: str) -> dict:
    """
    Extract internal hyperlinks from each page using PyMuPDF.
    Returns {page_num: [{text: str, dest_page: int}]}.

    In GRI content indices, disclosure codes are often hyperlinked to the
    page where the data is reported. We use these links to record the
    actual reporting page rather than the content index page.
    """
    result: dict = {}
    try:
        import fitz
        doc = fitz.open(pdf_path)
        for i in range(len(doc)):
            pg      = i + 1
            page    = doc[i]
            on_page = []
            for link in page.get_links():
                if link.get("kind") == 1:          # internal link
                    dest = link.get("page", -1)
                    if dest >= 0:
                        rect      = link.get("from")
                        link_text = page.get_text("text", clip=fitz.Rect(rect)).strip() if rect else ""
                        on_page.append({
                            "text":      link_text,
                            "dest_page": dest + 1,   # 0-indexed → 1-indexed
                        })
            if on_page:
                result[pg] = on_page
        doc.close()
        total_links = sum(len(v) for v in result.values())
        print(f"  [LINKS] {total_links} internal hyperlinks across {len(result)} pages")
    except Exception as e:
        print(f"  [WARN] Hyperlink extraction failed: {e}")
    return result


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 5 — ANALYSIS  (code-only matching)
# ══════════════════════════════════════════════════════════════════════════════

def _build_patterns(codes: dict) -> dict:
    return {
        code: re.compile(
            rf"(?:GRI\s+)?(?:Disclosure\s+)?{re.escape(code)}\b", re.I
        )
        for code in codes
    }


def find_disclosures_with_pages(pages: dict, hyperlinks: dict, codes: dict) -> dict:
    """
    Search every page for each GRI disclosure code.
    Returns {code: [page_nums]} for found codes only.

    When a code is found in the content index (detected via internal hyperlink
    pointing from the code text to a different page), the destination page is
    recorded instead of the index page — giving the actual reporting page.
    """
    patterns = _build_patterns(codes)
    found: dict[str, set] = {}

    for pg_num, page_data in pages.items():
        text = page_data["combined"]
        for code, pat in patterns.items():
            if not pat.search(text):
                continue
            # Check whether any hyperlink on this page originates near the code
            linked_dest = None
            for link in hyperlinks.get(pg_num, []):
                if link["text"] and pat.search(link["text"]):
                    linked_dest = link["dest_page"]
                    break
            target = linked_dest if (linked_dest and linked_dest != pg_num) else pg_num
            found.setdefault(code, set()).add(target)

    return {code: sorted(pgs) for code, pgs in found.items()}


def coverage_rating(found: int, expected: int) -> str:
    if expected == 0 or found == 0:
        return "Minimal"
    pct = found / expected
    if pct >= 0.75:
        return "Strong"
    if pct >= 0.50:
        return "Moderate"
    if pct > 0:
        return "Partial"
    return "Minimal"


def _std_display(sid: str, standards: dict) -> str:
    std = standards.get(sid)
    if std:
        return f"{std['standard_id']}: {std['name']} {std['version']}"
    return sid


def analyse_report(pages: dict, hyperlinks: dict, config: dict, standards: dict) -> dict:
    sector = config.get("sector_standard")

    results: dict = {
        "metadata": {
            "company_name":    config["company_name"],
            "report_year":     config["report_year"],
            "report_type":     config["report_type"],
            "sector_standard": sector,
            "total_pages":     len(pages),
            "extraction_ts":   datetime.datetime.now().isoformat(timespec="seconds"),
        },
        "gri_coverage": [],
    }

    for idx, (topic, cfg) in enumerate(GRI_TOPICS.items(), 1):
        print(f"  [ANALYSE] {topic}")

        expected = get_topic_disclosures(cfg, standards, sector)

        if not expected:
            results["gri_coverage"].append({
                "sr_no":                idx,
                "gri_topic":            topic,
                "underlying_standard":  "No GRI Standard",
                "coverage_rating":      "N/A",
                "pages":                "N/A",
                "disclosures":          "N/A",
                "reported":             False,
                "found_disclosures":    [],
                "expected_disclosures": [],
                "coverage_pct":         0,
            })
            continue

        found_map = find_disclosures_with_pages(pages, hyperlinks, expected)
        all_pages = sorted({p for pgs in found_map.values() for p in pgs})

        # Build sector-specific label lookup for display
        sector_labels: dict[str, str] = {}
        for sec_discs in cfg.get("sector_disclosures", {}).values():
            sector_labels.update(sec_discs)

        disc_strs = []
        for code in sorted(found_map.keys()):
            label = sector_labels.get(code, "")
            disc_strs.append(f"{code} ({label})" if label else code)

        # Underlying standard display names
        std_names = [_std_display(sid, standards) for sid in cfg.get("standard_ids", [])]
        if sector and cfg.get("sector_disclosures", {}).get(sector):
            std_names.append(f"{sector} (sector-specific)")
        if not std_names:
            std_names = ["No GRI Standard (sector-specific)"]

        results["gri_coverage"].append({
            "sr_no":                idx,
            "gri_topic":            topic,
            "underlying_standard":  " + ".join(std_names),
            "coverage_rating":      coverage_rating(len(found_map), len(expected)),
            "pages":                ", ".join(str(p) for p in all_pages) if all_pages else "N/A",
            "disclosures":          ", ".join(disc_strs) if disc_strs else "N/A",
            "reported":             len(found_map) > 0,
            "found_disclosures":    list(found_map.keys()),
            "expected_disclosures": list(expected.keys()),
            "coverage_pct":         round(len(found_map) / len(expected) * 100, 1),
        })

    return results


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 6 — JSON OUTPUT
# ══════════════════════════════════════════════════════════════════════════════

def save_json(results: dict, path: str):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print(f"  [OK] JSON → {path}")


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 7 — EXCEL OUTPUT  (single GRI Coverage sheet)
# ══════════════════════════════════════════════════════════════════════════════

def _hdr(ws, row: int, col: int, val: str, width: float | None = None):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(bold=True, size=10, name="Calibri")
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def _cell(ws, row: int, col: int, val, bold: bool = False, align: str = "left"):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(bold=bold, size=10, name="Calibri")
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal=align)
    return c


def _sheet_coverage(wb: Workbook, results: dict):
    ws   = wb.active
    ws.title = "GRI Coverage"
    meta = results["metadata"]
    rows = [r for r in results["gri_coverage"] if r.get("reported", False)]

    hdrs   = [
        "Sr. No",
        "GRI Topic",
        "Underlying GRI Standard",
        f"{meta['company_name']} Coverage",
        "Page No.",
        "Disclosure",
    ]
    widths = [8, 32, 52, 18, 25, 65]
    for col, (h, w) in enumerate(zip(hdrs, widths), 1):
        _hdr(ws, 1, col, h, width=w)
    ws.row_dimensions[1].height = 30

    for i, r in enumerate(rows, 1):
        row = i + 1
        _cell(ws, row, 1, i,                       align="center")
        _cell(ws, row, 2, r["gri_topic"],           bold=True)
        _cell(ws, row, 3, r["underlying_standard"])
        _cell(ws, row, 4, r["coverage_rating"],     align="center")
        _cell(ws, row, 5, r["pages"],               align="center")
        _cell(ws, row, 6, r["disclosures"])
        ws.row_dimensions[row].height = 35

    ws.freeze_panes = "A2"


def save_excel(results: dict, path: str):
    wb = Workbook()
    _sheet_coverage(wb, results)
    wb.save(path)
    print(f"  [OK] Excel → {path}")


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 8 — RUNNER
# ══════════════════════════════════════════════════════════════════════════════

def run(config: dict | None = None):
    if config is None:
        config = CONFIG

    os.makedirs(config["output_dir"], exist_ok=True)
    slug  = config["company_name"].replace(" ", "_")
    year  = config["report_year"]
    jpath = os.path.join(config["output_dir"], f"{slug}_{year}.json")
    xpath = os.path.join(config["output_dir"], f"{slug}_{year}.xlsx")

    banner = "=" * 64
    print(f"\n{banner}")
    print(f"  Sustainability Report Extractor")
    print(f"  Company : {config['company_name']}")
    print(f"  Year    : {config['report_year']}")
    print(f"  Sector  : {config.get('sector_standard') or 'None (topic standards only)'}")
    print(f"  PDF     : {Path(config['pdf_path']).name}")
    print(f"{banner}\n")

    print("[STEP 1/5]  Loading GRI standards corpus …")
    standards = load_gri_standards(_REPO_DIR)

    print("\n[STEP 2/5]  Extracting text from PDF …")
    pages = extract_text_by_page(config["pdf_path"])
    if not pages:
        print("[ERROR]  No text extracted — check PDF path and format.")
        return

    print("\n[STEP 3/5]  Extracting internal hyperlinks …")
    hyperlinks = extract_hyperlinks_by_page(config["pdf_path"])

    print(f"\n[STEP 4/5]  Matching {len(GRI_TOPICS)} GRI topics against disclosure codes …")
    results = analyse_report(pages, hyperlinks, config, standards)

    print("\n[STEP 5/5]  Saving outputs …")
    save_json(results, jpath)
    save_excel(results, xpath)

    print(f"\n{banner}")
    print("  COVERAGE SUMMARY")
    print(banner)
    tally: dict[str, list] = {"Strong": [], "Moderate": [], "Partial": [], "Minimal": []}
    not_reported = []
    for r in results["gri_coverage"]:
        if r.get("reported"):
            tally[r["coverage_rating"]].append(r["gri_topic"])
        else:
            not_reported.append(r["gri_topic"])

    for rating, topic_list in tally.items():
        if topic_list:
            print(f"\n  {rating} ({len(topic_list)})")
            for t in topic_list:
                print(f"    • {t}")
    if not_reported:
        print(f"\n  Not reported ({len(not_reported)})")
        for t in not_reported:
            print(f"    - {t}")

    reported_count = sum(len(v) for v in tally.values())
    print(f"\n  Topics reported : {reported_count} / {len(results['gri_coverage'])}")
    print(f"  Output folder   : {config['output_dir']}")
    print(f"  JSON            : {Path(jpath).name}")
    print(f"  Excel           : {Path(xpath).name}")
    print(banner)


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":

    _COMPANIES = _ROOT + "/pdfs/companies/"

    # ── Sony (no sector standard — topic standards only) ──────────────────────
    sony_config = {
        "company_name":    "Sony Group Corporation",
        "report_year":     "2025",
        "report_type":     "Sustainability Report",
        "sector_standard": None,
        "pdf_path":        _COMPANIES + "https::www.sony.com:en:SonyInfo:csr:library:reports:SustainabilityReport2025_E.pdf",
        "output_dir":      _ROOT + "/output",
    }

    # ── Vedanta (GRI 14: Mining sector) ──────────────────────────────────────
    vedanta_config = {
        "company_name":    "Vedanta Limited",
        "report_year":     "FY2024",
        "report_type":     "Sustainability Report",
        "sector_standard": "GRI 14",
        "pdf_path":        _COMPANIES + "https::www.vedantalimited.com:uploads:esg:esg-sustainability-framework:Sustainability-Report-FY2024.pdf",
        "output_dir":      _ROOT + "/output",
    }

    # ── Infosys (no sector standard) ──────────────────────────────────────────
    # infosys_config = {
    #     "company_name":    "Infosys Limited",
    #     "report_year":     "FY2025",
    #     "report_type":     "ESG Report",
    #     "sector_standard": None,
    #     "pdf_path":        _COMPANIES + "https::www.infosys.com:sustainability:documents:infosys-esg-report-2024-25.pdf",
    #     "output_dir":      _ROOT + "/output",
    # }

    # ── Wipro (no sector standard) ────────────────────────────────────────────
    # wipro_config = {
    #     "company_name":    "Wipro Limited",
    #     "report_year":     "FY2024",
    #     "report_type":     "Sustainability Report",
    #     "sector_standard": None,
    #     "pdf_path":        _COMPANIES + "https::www.wipro.com:content:dam:nexus:en:sustainability:sustainability_reports:wipro-sustainability-report-fy-2023-2024.pdf",
    #     "output_dir":      _ROOT + "/output",
    # }

    # ── MTAR (no sector standard) ─────────────────────────────────────────────
    # mtar_config = {
    #     "company_name":    "MTAR Technologies",
    #     "report_year":     "2026",
    #     "report_type":     "Annual Report",
    #     "sector_standard": None,
    #     "pdf_path":        _COMPANIES + "https::mtar.in:...",
    #     "output_dir":      _ROOT + "/output",
    # }

    run(sony_config)
