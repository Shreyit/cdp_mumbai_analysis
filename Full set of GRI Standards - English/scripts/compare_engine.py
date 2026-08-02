"""
================================================================================
  GRI Company Comparison Engine
================================================================================
  Loads two company JSONs (output of sustainability_extractor.py) and produces
  a structured comparison: per-topic scores, gap analysis, overall GRI score.

  Usage:
      python compare_engine.py  ← uses demo defaults below
      OR import and call compare(json_a, json_b)

  Output:
      output/comparison_<A>_vs_<B>.json
      output/comparison_<A>_vs_<B>.xlsx   (3-sheet comparison workbook)
================================================================================
"""

import os
import json
import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
RATING_SCORE = {"Strong": 3, "Moderate": 2, "Partial": 1, "Minimal": 0}
SCORE_RATING = {3: "Strong", 2: "Moderate", 1: "Partial", 0: "Minimal"}

_DARK_BLUE  = "1F3864"
_MED_BLUE   = "2E75B6"
_LIGHT_BLUE = "D6E4F0"
_WHITE      = "FFFFFF"
_GREY       = "F5F5F5"
_GREEN      = "C6EFCE"
_RED        = "FFC7CE"
_YELLOW     = "FFEB9C"

_thin   = Side(style="thin", color="CCCCCC")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

RATING_BG = {"Strong": "C6EFCE", "Moderate": "FFEB9C", "Partial": "FFCC99", "Minimal": "FFC7CE"}
RATING_FG = {"Strong": "375623", "Moderate": "7D6608", "Partial": "7D3C00", "Minimal": "9C0006"}

OUTPUT_DIR = Path(__file__).parent.parent / "output"

# ──────────────────────────────────────────────────────────────────────────────
# CORE COMPARISON LOGIC
# ──────────────────────────────────────────────────────────────────────────────

def _coverage_lookup(data: dict) -> dict:
    """Convert gri_coverage list → dict keyed by gri_topic for fast lookup."""
    return {r["gri_topic"]: r for r in data.get("gri_coverage", [])}


def compare(data_a: dict, data_b: dict) -> dict:
    """
    Compare two company report JSONs (gri_coverage format).
    Returns structured comparison dict with per-topic and aggregate metrics.
    """
    meta_a   = data_a["metadata"]
    meta_b   = data_b["metadata"]
    lookup_a = _coverage_lookup(data_a)
    lookup_b = _coverage_lookup(data_b)
    all_topics = list(dict.fromkeys(
        [r["gri_topic"] for r in data_a.get("gri_coverage", [])] +
        [r["gri_topic"] for r in data_b.get("gri_coverage", [])]
    ))

    topic_results = {}
    score_sum_a, score_sum_b, max_score = 0, 0, 0

    for topic in all_topics:
        ta = lookup_a.get(topic, {})
        tb = lookup_b.get(topic, {})

        rating_a = ta.get("coverage_rating", "Minimal")
        rating_b = tb.get("coverage_rating", "Minimal")
        score_a  = RATING_SCORE.get(rating_a, 0)
        score_b  = RATING_SCORE.get(rating_b, 0)

        # disclosures is now a comma-separated string; split for counting
        disc_a = [d.strip() for d in ta.get("disclosures", "N/A").split(",") if d.strip() != "N/A"]
        disc_b = [d.strip() for d in tb.get("disclosures", "N/A").split(",") if d.strip() != "N/A"]
        exp    = ta.get("expected_disclosures", tb.get("expected_disclosures", []))
        exp_n  = max(len(exp), 1)

        gap    = score_a - score_b
        winner = meta_a["company_name"] if gap > 0 else (
                 meta_b["company_name"] if gap < 0 else "Tied")

        score_sum_a += score_a
        score_sum_b += score_b
        max_score   += 3

        topic_results[topic] = {
            "company_a": {
                "name":                    meta_a["company_name"],
                "rating":                  rating_a,
                "score":                   score_a,
                "page_range":              ta.get("page_range", "N/A"),
                "disclosures":             ta.get("disclosures", "N/A"),
                "disclosure_coverage_pct": round(len(disc_a) / exp_n * 100, 1),
            },
            "company_b": {
                "name":                    meta_b["company_name"],
                "rating":                  rating_b,
                "score":                   score_b,
                "page_range":              tb.get("page_range", "N/A"),
                "disclosures":             tb.get("disclosures", "N/A"),
                "disclosure_coverage_pct": round(len(disc_b) / exp_n * 100, 1),
            },
            "gap":    gap,
            "winner": winner,
        }

    overall_a = round(score_sum_a / max(max_score, 1) * 100, 1)
    overall_b = round(score_sum_b / max(max_score, 1) * 100, 1)

    # Breakdown by pillar
    pillars = {
        "Environment":  ["Climate Change", "Air Emissions", "Biodiversity", "Waste", "Water and Effluents"],
        "Social":       ["Occupational Health & Safety", "Employment Practices", "Child Labour",
                         "Forced Labour & Modern Slavery", "Freedom of Association",
                         "Non-discrimination & Equal Opportunity", "Human Rights",
                         "Local Communities", "Security Practices"],
        "Governance":   ["Anti-corruption", "Payments to Governments", "Public Policy",
                         "Corporate Governance"],
        "Economic":     ["Economic Impacts"],
        "Other":        ["Responsible Supply Chain", "Product Safety & Quality",
                         "Data Privacy & Cybersecurity", "Responsible AI"],
    }
    pillar_scores = {}
    for pillar, p_topics in pillars.items():
        pa, pb, pm = 0, 0, 0
        for t in p_topics:
            if t in topic_results:
                pa += topic_results[t]["company_a"]["score"]
                pb += topic_results[t]["company_b"]["score"]
                pm += 3
        if pm > 0:
            pillar_scores[pillar] = {
                "company_a_pct": round(pa / pm * 100, 1),
                "company_b_pct": round(pb / pm * 100, 1),
            }

    return {
        "metadata": {
            "company_a":      meta_a["company_name"],
            "company_b":      meta_b["company_name"],
            "year_a":         meta_a["report_year"],
            "year_b":         meta_b["report_year"],
            "compared_at":    datetime.datetime.now().isoformat(timespec="seconds"),
            "total_topics":   len(topic_results),
        },
        "summary": {
            "overall_gri_score_a":   overall_a,
            "overall_gri_score_b":   overall_b,
            "overall_winner":        meta_a["company_name"] if overall_a > overall_b else (
                                     meta_b["company_name"] if overall_b > overall_a else "Tied"),
            "topics_a_leads":        sum(1 for t in topic_results.values() if t["gap"] > 0),
            "topics_b_leads":        sum(1 for t in topic_results.values() if t["gap"] < 0),
            "topics_tied":           sum(1 for t in topic_results.values() if t["gap"] == 0),
            "pillar_scores":         pillar_scores,
        },
        "topics": topic_results,
    }


# ──────────────────────────────────────────────────────────────────────────────
# EXCEL OUTPUT
# ──────────────────────────────────────────────────────────────────────────────

def _hdr(ws, row, col, val, bg=_DARK_BLUE, fg=_WHITE, sz=10, bold=True, center=True):
    c = ws.cell(row=row, column=col, value=val)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.font      = Font(color=fg, bold=bold, size=sz, name="Calibri")
    c.alignment = Alignment(wrap_text=True, vertical="center",
                            horizontal="center" if center else "left")
    c.border    = _border
    return c

def _dat(ws, row, col, val, bg=_WHITE, bold=False, align="left", sz=10, fg="000000"):
    c = ws.cell(row=row, column=col, value=val)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.font      = Font(bold=bold, size=sz, name="Calibri", color=fg)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal=align)
    c.border    = _border
    return c

def _title_row(ws, text, ncols, row=1, bg=_DARK_BLUE, sz=13):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.font      = Font(color=_WHITE, bold=True, size=sz, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30

def _rating_cell(ws, row, col, rating, bg_override=None):
    bg = bg_override or RATING_BG.get(rating, _WHITE)
    fg = RATING_FG.get(rating, "000000")
    c  = _dat(ws, row, col, rating, bg=bg, bold=True, align="center", fg=fg)
    return c


def _sheet_company(wb, data: dict, first: bool = False):
    """
    One sheet per company — exact same 6-column format as GRI14_Vedanta_Coverage_Mapping.csv.
    Sheet name = "CompanyName (Year)"  e.g. "Sony Group (2025)"
    """
    meta  = data["metadata"]
    rows  = data["gri_coverage"]
    short = meta["company_name"].split()[0]          # first word for short sheet name
    title = f"{short} ({meta['report_year']})"

    ws = wb.active if first else wb.create_sheet(title)
    ws.title = title
    ws.sheet_view.showGridLines = False

    _title_row(ws,
        f"{meta['company_name']} — {meta['report_type']} {meta['report_year']} | GRI Coverage Mapping",
        ncols=6, row=1)

    ws.merge_cells("A2:F2")
    sub = ws.cell(row=2, column=1,
        value=f"Pages: {meta['total_pages']}   |   Extracted: {meta['extraction_ts'][:10]}")
    sub.fill      = PatternFill("solid", fgColor=_MED_BLUE)
    sub.font      = Font(color=_WHITE, size=10, name="Calibri")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    hdrs   = ["Sr. No", "GRI Topic", "Underlying GRI Topic Standard",
              f"{meta['company_name']} Coverage", "Page No.", "Disclosure"]
    widths = [8, 34, 54, 18, 16, 50]
    for col, (h, w) in enumerate(zip(hdrs, widths), 1):
        _hdr(ws, 3, col, h, bg=_MED_BLUE)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 34

    for r in rows:
        row    = r["sr_no"] + 3
        bg     = _GREY if r["sr_no"] % 2 == 0 else _WHITE
        rating = r["coverage_rating"]
        r_bg   = RATING_BG.get(rating, _WHITE)
        r_fg   = RATING_FG.get(rating, "000000")

        _dat(ws, row, 1, r["sr_no"],              bg=bg, align="center")
        _dat(ws, row, 2, r["gri_topic"],           bg=bg, bold=True)
        _dat(ws, row, 3, r["underlying_standard"], bg=bg)
        c = _dat(ws, row, 4, rating, bg=r_bg, bold=True, align="center")
        c.font = Font(bold=True, size=10, color=r_fg, name="Calibri")
        _dat(ws, row, 5, r["page_range"],  bg=bg, align="center")
        _dat(ws, row, 6, r["disclosures"], bg=bg)
        ws.row_dimensions[row].height = 40

    ws.freeze_panes = "A4"


def _sheet_comparison(wb, datasets: list):
    """
    Side-by-side gap sheet — one column per company for rating + winner column.
    Only added when 2+ companies are compared.
    """
    ws = wb.create_sheet("Comparison")
    ws.sheet_view.showGridLines = False

    names  = [d["metadata"]["company_name"] for d in datasets]
    years  = [d["metadata"]["report_year"]  for d in datasets]
    labels = [f"{n} ({y})" for n, y in zip(names, years)]
    ncols  = 2 + len(datasets) * 2 + 1          # Sr + Topic + (Rating+Pages)*N + Winner

    _title_row(ws, "GRI Coverage — Side-by-Side Comparison", ncols=ncols, row=1)

    # Header row
    hdrs = ["Sr. No", "GRI Topic"]
    for lbl in labels:
        hdrs += [f"{lbl}\nRating", f"{lbl}\nPage No."]
    hdrs.append("Leader")
    widths = [8, 34] + [18, 14] * len(datasets) + [28]

    for col, (h, w) in enumerate(zip(hdrs, widths), 1):
        _hdr(ws, 2, col, h, bg=_MED_BLUE)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 40

    # Build topic lookup: topic → [row_dict per company]
    topic_order = [r["gri_topic"] for r in datasets[0]["gri_coverage"]]
    lookup = {
        d["metadata"]["company_name"]: {r["gri_topic"]: r for r in d["gri_coverage"]}
        for d in datasets
    }

    for idx, topic in enumerate(topic_order, 1):
        row = idx + 2
        bg  = _GREY if idx % 2 == 0 else _WHITE

        _dat(ws, row, 1, idx, bg=bg, align="center")
        _dat(ws, row, 2, topic, bg=bg, bold=True)

        scores = []
        col = 3
        for name in names:
            r      = lookup[name].get(topic, {})
            rating = r.get("coverage_rating", "Minimal")
            pages  = r.get("page_range", "N/A")
            r_bg   = RATING_BG.get(rating, _WHITE)
            r_fg   = RATING_FG.get(rating, "000000")
            c = _dat(ws, row, col, rating, bg=r_bg, bold=True, align="center")
            c.font = Font(bold=True, size=10, color=r_fg, name="Calibri")
            _dat(ws, row, col + 1, pages, bg=bg, align="center")
            scores.append((name, RATING_SCORE.get(rating, 0)))
            col += 2

        # Winner column
        max_score = max(s for _, s in scores)
        leaders   = [n for n, s in scores if s == max_score]
        leader    = leaders[0] if len(leaders) == 1 else "Tied"
        lead_bg   = _GREEN if len(leaders) == 1 else _YELLOW
        _dat(ws, row, col, leader, bg=lead_bg, bold=True, align="center")
        ws.row_dimensions[row].height = 40

    ws.freeze_panes = "A3"


def save_comparison_excel(datasets: list, path: str):
    """
    Build workbook: one sheet per company (CSV format) + Comparison sheet.
    datasets = list of company JSON dicts (gri_coverage format).
    """
    wb = Workbook()
    for i, data in enumerate(datasets):
        _sheet_company(wb, data, first=(i == 0))
    if len(datasets) >= 2:
        _sheet_comparison(wb, datasets)
    wb.save(path)
    print(f"  [OK] Excel → {path}")


# ──────────────────────────────────────────────────────────────────────────────
# RUNNER  — accepts any number of company JSON paths
# ──────────────────────────────────────────────────────────────────────────────

def run(*paths: str, output_dir: str = None):
    """
    Pass 2 or more paths to company JSONs.
    Produces one sheet per company + a Comparison sheet.
    """
    if len(paths) < 2:
        print("[ERROR] Provide at least 2 company JSON paths to compare.")
        return

    out = Path(output_dir or OUTPUT_DIR)
    out.mkdir(parents=True, exist_ok=True)

    datasets = []
    print(f"\n{'='*60}\n  Comparing:\n{'='*60}")
    for i, p in enumerate(paths, 1):
        with open(p, encoding="utf-8") as f:
            d = json.load(f)
        datasets.append(d)
        print(f"  {i}. {d['metadata']['company_name']} ({d['metadata']['report_year']})")

    slugs = "_vs_".join(
        d["metadata"]["company_name"].replace(" ", "_") for d in datasets
    )
    base       = f"comparison_{slugs}"
    excel_path = out / f"{base}.xlsx"

    save_comparison_excel(datasets, str(excel_path))

    # Console summary (pairwise for first 2)
    if len(datasets) == 2:
        result = compare(datasets[0], datasets[1])
        s = result["summary"]
        print(f"\n  Overall GRI Score:")
        print(f"    {datasets[0]['metadata']['company_name']}: {s['overall_gri_score_a']}%")
        print(f"    {datasets[1]['metadata']['company_name']}: {s['overall_gri_score_b']}%")
        print(f"  Winner : {s['overall_winner']}")
        print(f"  Leads  : {datasets[0]['metadata']['company_name']} {s['topics_a_leads']}  |  "
              f"{datasets[1]['metadata']['company_name']} {s['topics_b_leads']}  |  "
              f"Tied {s['topics_tied']}")

    print(f"\n  Excel → {excel_path}\n{'='*60}")


if __name__ == "__main__":
    # Add as many company JSON paths as needed — one sheet per company is created
    paths = [
        str(OUTPUT_DIR / "Sony_Group_Corporation_2025.json"),
        str(OUTPUT_DIR / "Vedanta_Limited_FY2024.json"),
        # str(OUTPUT_DIR / "Infosys_Limited_FY2025.json"),   # uncomment to add 3rd company
        # str(OUTPUT_DIR / "Wipro_Limited_FY2024.json"),     # uncomment to add 4th company
    ]
    existing = [p for p in paths if Path(p).exists()]
    if len(existing) >= 2:
        run(*existing)
    else:
        print("Run sustainability_extractor.py for at least 2 companies first.")
        print("Expected files in output/:")
        for p in paths:
            status = "✓" if Path(p).exists() else "✗ missing"
            print(f"  {status}  {Path(p).name}")
