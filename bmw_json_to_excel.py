"""
BMW Group CDP corpus -> single Excel workbook.
==============================================
Clone of apple_json_to_excel.py for the BMW Group CDP responses (2021-2025).

Reads the per-year JSON files produced by bmw_pdf_to_json.py and writes one
long-format workbook covering all five years.

Columns (same schema as Apple_Responses.xlsx):
  str      - question code, plain numbering (0.1, 1.1, 7.6.1, ...; legacy
             C-codes have the "C" stripped so C6.1 -> 6.1; the compound
             C-TO0.7/C-TS0.7 is kept verbatim)
  year     - filing year (2021 .. 2025)
  parent   - nearest two-part question id ("" for top-level two-part codes)
  section  - module header the question lives in (C0/C1/...)
  row_label- repeating-table category (new format only; "" otherwise)
  question - question / sub-field wording
  response - the answer text

Run:  .venv/bin/python bmw_json_to_excel.py
Output: BMW_GROUP/BMW_Responses.xlsx
"""

import json
from pathlib import Path

import pandas as pd

EXCEL_CELL_LIMIT = 32767  # hard limit per cell in .xlsx
TRUNC_SUFFIX = "\n[…truncated…]"

BMW_DIR = Path(__file__).resolve().parent / "BMW_GROUP"
YEARS = [2021, 2022, 2023, 2024, 2025]
OUT = BMW_DIR / "BMW_Responses.xlsx"


def build() -> pd.DataFrame:
    rows = []
    for year in YEARS:
        src = BMW_DIR / f"BMW_CDP_{year}.json"
        data = json.loads(src.read_text(encoding="utf-8"))
        for q in data["questions"]:
            resp = q.get("response", "")
            if len(resp) > EXCEL_CELL_LIMIT:
                resp = resp[: EXCEL_CELL_LIMIT - len(TRUNC_SUFFIX)] + TRUNC_SUFFIX
            rows.append({
                "str": q["id"],
                "year": year,
                "parent": q.get("parent", ""),
                "section": q.get("section", ""),
                "row_label": q.get("row_label", ""),
                "question": q["question"],
                "response": resp,
            })
    return pd.DataFrame(rows)


def main():
    df = build()
    from openpyxl.styles import Font, PatternFill
    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="BMW Responses", index=False)
        ws = writer.sheets["BMW Responses"]
        widths = {"A": 12, "B": 8, "C": 10, "D": 28, "E": 40, "F": 70, "G": 90}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    print(f"wrote {OUT} with {len(df)} rows")
    print(df.groupby("year").size().to_string())


if __name__ == "__main__":
    main()
